from openpyxl import Workbook

wb = Workbook()
# wb.remove_sheet()
wb.create_sheet('Shit1')
ws = wb.active
print(ws.get_named_range)
# ws['A1'] = 42
# ws.append([1, 2, 3])
# wb.save('sample.xlsx')
print(wb.sheetnames)